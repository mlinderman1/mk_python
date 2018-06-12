import time,re,json,math,os,httplib,shutil
from openpyxl import Workbook
from openpyxl.styles import Font
from urllib import unquote

from selenium import webdriver
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common import exceptions as driverExceptions
from selenium.webdriver.common.keys import Keys
import xml.etree.ElementTree as ET
import Tkinter as tk
import tkFileDialog

webdriver;

global debug
debug=[]
#ToDo: Classes
#ToDo: Payment options for DD / Open Invoice

def filterNetworkLog(log):
    #Takes entire Chrome Driver network log and pulls out relevant entries
    #Ignores entries caused by cross country consistency actions embedded in the script
    #e.g. automatic colour changes on PLPs
    requestArray = []
    for i in log:
        messageTemp = json.loads(i['message'])
        if messageTemp['message']['method']=='Network.requestWillBeSent':
            if messageTemp['message']['params']['request']['method']=='GET' and "michaelkors.d2.sc.omtrdc.net/b/ss" in messageTemp['message']['params']['request']['url']:
                main = re.match('.+\?(.*)',messageTemp['message']['params']['request']['url']).groups()[0]
                rs = re.match('.*/b/ss/([^/]+)',messageTemp['message']['params']['request']['url']).groups()[0]
                if 'IgnoreFlagDoNotInclude' not in main:
                    requestArray.append('rs='+rs+'&'+main)
            elif messageTemp['message']['params']['request']['method']=='POST' and "michaelkors.d2.sc.omtrdc.net/b/ss" in messageTemp['message']['params']['request']['url']:
                main = messageTemp['message']['params']['request']['postData']
                rs = re.match('.*/b/ss/([^/]+)',messageTemp['message']['params']['request']['url']).groups()[0]
                if 'IgnoreFlagDoNotInclude' not in main:
                    requestArray.append('rs='+rs+'&'+main)
    return requestArray

def parseSeleniumScript(file):
    # Parses Katalon recorder .xml file as executable in this script

    tree = ET.parse(file)
    root = tree.getroot()

    actionTargetValue = []
    toKeep = ['selectFrame','click','submit','type','sendKeys','select','mouseOver']

    for i in root:
        if i.find('command').text in toKeep:

            if i.find('command').text=='type' or i.find('command').text=='select':
                value = i.find('value').text
            elif i.find('command').text=='sendKeys':
                key = re.match('.*_([^}]+)',i.find('value').text).groups()[0]
                value = eval('Keys.'+key)
            else:
                value=''

            if i.find('target').text[:2]=='//' or i.find('target').text[:5]=='xpath':
                finder = 'xpath'
                target = re.match('(xpath=)?(^\/\/.*|.*)',i.find('target').text).groups()[1]
            elif i.find('target').text[:5]=='link=':
                finder = 'linkText'
                target = re.sub('link=','',i.find('target').text)
            else:
                finder = 'css'
                if i.find('target').text[:3]=='css':
                    target = re.sub('\.( |$)',' ',i.find('target').text[4:])
                elif i.find('command').text=='selectFrame':
                    target = re.sub('.*=','',i.find('target').text)
                    try:
                        target = int(target)
                    except ValueError:
                        pass
                else:
                    target = re.match('([^=]+=)(.*)',re.sub('\.( |$)',' ',i.find('target').text)).groups()
                    target = '['+target[0]+'"'+target[1]+'"]'

            actionTargetValue.append({'action':i.find('command').text,'target':target,'value':value,'finder':finder})
    return actionTargetValue

def doDL(*andReturn):
    #Sets JS script up to poll datalayer every .2 seconds and save in browser sessionStorage when changed
    #Returns all captured DLs when instructed
    dataLayerScript = "return(function(w){w.removeListing=w.removeListing||function(dl){var d=JSON.parse(JSON.stringify(dl));if(d.page&&d.page.listing){d.page.listing='...'};if(d.product&&d.product.SkuVariants){d.product.SkuVariants='...'};if(d.event){for(i=0;i<d.event.length;i++){if(d.event[i].page&&d.event[i].page.listing){d.event[i].page.listing='...'};if(d.event[i].product&&d.event[i].product.SkuVariants){d.event[i].product.SkuVariants='...'}}};return d};w.arrayDL=w.arrayDL||{};w.doCheck=function(){var nd=new Date();if(!w.sessionStorage.count||w.sessionStorage.dl_last!=JSON.stringify(w.removeListing(w.mkorsData))){w.sessionStorage.setItem('dl_last',JSON.stringify(w.removeListing(w.mkorsData)));w.sessionStorage.setItem('count',parseInt((w.sessionStorage.count||0))+1);w.sessionStorage.setItem('dl_'+w.sessionStorage.count,JSON.stringify({'dl':w.removeListing(w.mkorsData),'url':w.location.href}));for(var i=1;i<parseInt(w.sessionStorage.count)+1;i++){w.arrayDL[i-1]=JSON.parse(w.sessionStorage['dl_'+i])}};w.sessionStorage.setItem('tdl_'+w.sessionStorage.count,nd.toTimeString().substring(0,8))};w.doCheck();w.update=w.update||setInterval(w.doCheck,200);for(i in w.arrayDL){w.arrayDL[i]['t']=w.sessionStorage['tdl_'+(1*i+1)]};return JSON.stringify(w.arrayDL)})(window.top)"
    jsDL = driver.execute_script(dataLayerScript)

    if andReturn:
        jsDLS = json.loads(jsDL)
        currentLayer={}
        for DLs in jsDLS:
            currentLayer[jsDLS[DLs]['t']]={'dl':jsDLS[DLs]['dl'],'url':jsDLS[DLs]['url']}
        return currentLayer

def seleniumMain(script,site,replacedElems,sitesList,device):
    #Initialises ChromeDriver
    findCountry = 0
    oldURL = oldQB = ''
    siteDomain = re.match('.*\.michaelkors\.([^/]+(/en_(PL|NL|HU))?)',site).groups()[0]

    if siteDomain=='co.uk':
        countryPostal = 'WC2B 6UF'

    elif (siteDomain == 'ch' or siteDomain=='eu/en_HU'):
        countryPostal = '1234'

    elif siteDomain == 'eu/en_NL':
        countryPostal = '1234 AA'

    elif siteDomain == 'eu/en_PL':
        countryPostal = '12-123'

    else:
        countryPostal = '12345'

    if 'driver' not in globals():
        global driver
        pathToChromeTemp = "chromedriver"
        # pathToChromeTemp = "C:\\Users\\" + os.environ.get('USERNAME')+"\\Documents\\ChromeTest"
        try:
            shutil.rmtree(pathToChromeTemp)
        except Exception:
            pass
        capabilities = DesiredCapabilities.CHROME
        capabilities['loggingPrefs'] = {'performance': 'INFO'}
        options = webdriver.ChromeOptions()
        if device == 'desktop':
            options.add_argument("--start-maximized")
        elif device == 'mobile':
            ua = "Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1"
            options.add_argument("user-agent=" + ua)
        elif device == 'tablet':
            ua = "Mozilla/5.0 (iPad; CPU OS 11_0 like Mac OS X) AppleWebKit/604.1.34 (KHTML, like Gecko) Version/11.0 Mobile/15A5341f Safari/604.1"
            options.add_argument("user-agent=" + ua)
        options.add_argument("--disable-web-security")
        options.add_argument("--allow-running-insecure-content")
        options.add_argument("--user-data-dir="+pathToChromeTemp)
        options.add_argument("--proxy-server=10.143.82.20:8080")
        # Chrome webdriver will need occasional updating, depending on your version of Chrome auto-updating
        # Details at https://sites.google.com/a/chromium.org/chromedriver/downloads

        driver = webdriver.Chrome(executable_path='chromedriver',desired_capabilities=capabilities,options=options)
        # driver = webdriver.Chrome(executable_path='G:\\eCommerce Europe\\BAU\\Analytics & Reporting\\6- Tools\\Selenium\\Chrome\\chromedriver.exe',desired_capabilities=capabilities,options=options)

    sizes = {'tablet':{'x':1024,'y':768},'mobile':{'x':414,'y':736}}

    if device!='desktop':
        driver.set_window_size(sizes[device]['x'],sizes[device]['y'])

    driver.delete_all_cookies()

    #Setting 'gioIp = renderCookie' cookie prevents country redirect modal appearing on visiting non-UK site
    domain = re.match('.*//([^/]+)',site).groups()[0]
    driver.add_cookie({'name' : 'gioIp', 'value' : 'renderCookie', 'path' : '/','domain' : domain})

    driver.get(site)
    driver.execute_script('sessionStorage.clear()')

    first = 1 if siteDomain==sitesList[0] else 0

    for i in script:

        doDL()
        debug.append(json.dumps(i))

        print i
        last = i

        if i['target']=='div.ajax_overlay' and '/checkout/checkout.jsp' in driver.current_url:
            continue

        if i['action']=='selectFrame':
            try:
                doFrameSteps(i)
            except httplib.CannotSendRequest:
                #Bug between selenium and httplib means CannotSendRequest is sometimes thrown
                #when initially accessing iframe elements, but works fine 2nd time
                doFrameSteps(i)
        else:
            try:
                olds = doSteps(i,findCountry,siteDomain,countryPostal,oldURL,oldQB,replacedElems,first)
                oldURL = olds[0]
                oldQB = olds[1]
                replacedElems = olds[2]
                findCountry = olds[3]
            except httplib.CannotSendRequest:
                # Bug between selenium and httplib means CannotSendRequest is sometimes thrown
                # when initially accessing iframe elements, but works fine 2nd time
                olds = doSteps(i,findCountry,siteDomain,countryPostal,oldURL,oldQB,replacedElems,first)
                oldURL = olds[0]
                oldQB = olds[1]
                replacedElems = olds[2]
                findCountry = olds[3]
            except driverExceptions.TimeoutException:
                # If attempts to find current element time out, there is a good chance this is because
                # The previous step failed (e.g. didn't click link to expand an element)
                # So repeats previous step
                olds = doSteps(last,findCountry,siteDomain,countryPostal,oldURL,oldQB,replacedElems,first)
                oldURL = olds[0]
                oldQB = olds[1]
                replacedElems = olds[2]
                findCountry = olds[3]

                olds = doSteps(i,findCountry,siteDomain,countryPostal,oldURL,oldQB,replacedElems,first)
                oldURL = olds[0]
                oldQB = olds[1]
                replacedElems = olds[2]
                findCountry = olds[3]
            except driverExceptions.StaleElementReferenceException:
                # StaleElementReferenceException is raised if DOM is
                # refreshed between element being located and actioned on
                olds = doSteps(i,findCountry,siteDomain,countryPostal,oldURL,oldQB,replacedElems,first)
                oldURL = olds[0]
                oldQB = olds[1]
                replacedElems = olds[2]
                findCountry = olds[3]
            except Exception,e:
                if 'stale element' in repr(e):
                    # Also raised in this format, due to other error catching in visibleClick (?)
                    olds = doSteps(i,findCountry,siteDomain,countryPostal,oldURL,oldQB,replacedElems,first)
                    oldURL = olds[0]
                    oldQB = olds[1]
                    replacedElems = olds[2]
                    findCountry = olds[3]
                else:
                    raise e

            last = i

    time.sleep(10)

    waitForAnalytics(oldURL,oldQB,i,1)


def doFrameSteps(i):
    # Actions inside iFrames are referenced by iFrame index number during the recording process
    # Indexes are sensitive to the order the iFrame is loaded, so may not match when running automation
    # This function assumes the iFrame to be interacted with is the only one in the DOM that is visible
    # Assumption is true for checkout(3ds),truefit & live chat
    if type(i['target'])==int:
        time.sleep(5)
        noDisplay = 1
        while noDisplay:
            for frames in driver.find_elements_by_tag_name('iframe'):
                if frames.is_displayed():
                    driver.switch_to_frame(frames)
                    noDisplay = 0
                    break
    else:
        driver.switch_to_default_content()

def doSteps(i,findCountry,siteDomain,countryPostal,oldURL,oldQB,replacedElems,isFirstCountry):

    def waitForElementAndSelect(current,replacedElems,oldURL,oldQB):
        # Locates element on page and waits until interactable

        elemRef = '_'+current['target']+re.sub('^\/?en_[A-Z][A-Z]','',driver.current_url.split('/').pop())
        newi = {}

        if current['action']=='select' and not current['target']=='[id="returnCode"]':
            # 'select' actions not needed anywhere but 3DS payment page
            newi['target'] = ''
            newi['finder'] = ''
            newi['value'] = ''
            newi['action'] = ''
            elem = ''

        elif isFirstCountry:
            # Grabs more specific element references in first country run through
            # As the initial references do not always work across further countries
            if current['finder']=='xpath':
                if "contains(@src" in current['target']:
                    # Strip hostname from src matches
                    newi['target'] = re.sub('http.*michaelkors\.[a-z]{2}(\.[a-z]{2})?','',current['target'])
                    newi['finder'] = 'xpath'
                    newi['value'] = current['value']
                    newi['action'] = current['action']
                    condition = EC.presence_of_element_located((By.XPATH,newi['target']+'[1]'))
                else:
                    condition = EC.presence_of_element_located((By.XPATH,current['target']+'[1]'))
            elif current['finder']=='css':
                condition = EC.presence_of_element_located((By.CSS_SELECTOR,current['target']))
            elif current['finder']=='linkText':
                #By.LINK_TEXT behaves inconsistently - xpath workaround
                newi['finder'] = 'xpath'
                newi['target'] = "(//a[normalize-space(text()) = '"+current['target']+"']|//a/*[normalize-space(text()) = '"+current['target']+"'])[1]"
                newi['value'] = current['value']
                newi['action'] = current['action']
                condition =  EC.presence_of_element_located((By.XPATH,newi['target']))

            elem = WebDriverWait(driver, 15).until(condition)

            olds = waitForAnalytics(oldURL,oldQB,newi or current)
            oldURL = olds[0]
            oldQB = olds[1]

            if elem.find_elements_by_xpath('ancestor::*[contains(@class, "product-tile")]'):
                # Products are displayed in different order on cross country PLPs,
                # Locates quickview buttons and colour changes by parent href instead of index
                # Only runs in recorded journey (always the first to run in automation)
                # creates lookup dict for further countries to reference
                # Captures default colour of products - not always the same cross country and can cause errors

                productTile = elem.find_element_by_xpath('ancestor::*[contains(@class, "product-tile")]')
                theHREF = productTile.find_element_by_xpath('.//div[@class="image-panel"]/a').get_property('pathname').split('/').pop()
                colourElem = elem.find_elements_by_xpath('ancestor::*[contains(@class, "product-tile")]//*[@class="swatch-link selected"]/img')

                initialColour = colourElem[0].get_attribute('data-skuimg') if colourElem else ''
                theColour = colourElem[0].get_attribute('title') if colourElem else ''
                replacedElems[elemRef] = {'target':'','initialColour':{'ref':initialColour,'colour':theColour}}

                if 'quickview-btn' in current['target'] or "button[@name='quickview-btn']" in current['target']:
                    replacedElems[elemRef]['target'] = 'a[href*="' + theHREF + '"] [name="quickview-btn"]'

                elif 'swatch-image' in elem.get_attribute('class'):
                    colour = elem.get_attribute('data-skuimg')
                    replacedElems[elemRef]['target'] = ' img[data-skuimg="' + colour + '"]'

                else:
                    replacedElems[elemRef]['target'] = '.product-tile a[href*="' + theHREF + '"]'

            elif elem.get_attribute('data-value'):
                # Edit product size selection on checkout pages matches by text e.g. UK 7
                # Match by unique 'data-value' attribute instead
                replacedElems[elemRef] = {'target':'[data-value="'+elem.get_attribute('data-value')+'"]','initialColour':{'ref':'','colour':''}}

            elif elem.find_elements_by_xpath('ancestor::li[contains(@class, "facet-color-options")]'):
                # Quickview colour options
                clickCol = elem.find_element_by_xpath('ancestor::li[contains(@class, "facet-color-options")]').get_attribute('title')
                replacedElems[elemRef] = {'target':'li.facet-color-options[title="'+clickCol+'"] label','initialColour':{'ref':'','colour':''}}

            elif '/checkout/checkout.jsp' in driver.current_url and current['finder']!='css' and elem.find_elements_by_xpath('ancestor::form[contains(@id, "payment-form")]'):
                # Checkout payment options - DE usually fails due to extra options (Open Invoice etc..)
                dataID = elem.get_attribute('data-id') or elem.find_element_by_xpath('..').get_attribute('data-id')
                container = elem.find_elements_by_xpath('ancestor::div[contains(@class,"payment-panel-body")]')
                currentTargetValue = current['target']
                if container:
                    if dataID:
                        # Month/Year dropdown
                        target = '#' + container[0].get_attribute('id') + ' [data-id='+dataID+']'
                        replacedElems[elemRef] = {'target':target,'initialColour':{'ref':'','colour':''}}
                    elif current['finder']=='linkText':
                        # Month/Year dropdown values
                        target = '(//div[@id="'+container[0].get_attribute('id')+'"]//span[normalize-space(text())="'+current['target']+'"])[1]'
                        replacedElems[elemRef] = {'target':target,'initialColour':{'ref':'','colour':''},'xpath':True}
            else:
                # Some text matches fail cross countries e.g. 'trousers' become 'pants'
                # Matches by parent elements and URL path instead
                translate =  ['trousers','trainers','jewellery','sunglasses','gifts under','purses']
                for i in translate:
                    if re.search(i,current['target'],flags=re.I):
                        theHREF = re.sub('^\/en_[A-Z][A-Z]','',elem.get_property('pathname')).split('/')
                        href0 = theHREF.pop()
                        href1 = theHREF.pop()
                        ancestors = elem.find_element_by_xpath('../../..').tag_name
                        ancestors+= ' > ' + elem.find_element_by_xpath('../..').tag_name
                        ancestors+= ' > ' + elem.find_element_by_xpath('..').tag_name
                        target = ancestors + ' > a[href$="/'+href1+'/'+href0+'"]'
                        replacedElems[elemRef] = {'target':target,'initialColour':{'ref':'','colour':''}}

        else:
            if elemRef in replacedElems and replacedElems[elemRef]['target']:
                # If current element exists in list of elements to translate, use that reference instead
                # driver.execute_script('window.scrollTo(0,document.body.scrollHeight)') # Loads more products, PLP lazy load
                if 'xpath' in replacedElems[elemRef]:
                    condition = EC.presence_of_element_located((By.XPATH,replacedElems[elemRef]['target']))
                else:
                    condition = EC.presence_of_element_located((By.CSS_SELECTOR,replacedElems[elemRef]['target']))
                newi['target'] = replacedElems[elemRef]['target']
                newi['finder'] = 'css'
                newi['value'] = current['value']
                newi['action'] = current['action']

            elif current['finder']=='xpath':
                if "contains(@src" in current['target']:
                    # Strip hostname from src matches
                    newi['target'] = re.sub('http.*michaelkors\.[a-z]{2}(\.[a-z]{2})?','',current['target'])
                    newi['finder'] = 'xpath'
                    newi['value'] = current['value']
                    newi['action'] = current['action']
                    condition = EC.presence_of_element_located((By.XPATH,newi['target']+'[1]'))
                else:
                    condition = EC.presence_of_element_located((By.XPATH,current['target']+'[1]'))
            elif current['finder']=='css':
                condition = EC.presence_of_element_located((By.CSS_SELECTOR,current['target']))
            elif current['finder']=='linkText':
                #By.LINK_TEXT behaves inconsistently - xpath workaround
                current['finder']='xpath'
                current['target']="(//a[normalize-space(text()) = '"+current['target']+"']|//a/*[normalize-space(text()) = '"+current['target']+"'])[1]"
                condition =  EC.presence_of_element_located((By.XPATH,current['target']))

            elem = WebDriverWait(driver, 15).until(condition)

            olds = waitForAnalytics(oldURL,oldQB,newi or current)
            oldURL = olds[0]
            oldQB = olds[1]

            productTile = elem.find_elements_by_xpath('ancestor::*[contains(@class, "product-tile")]')
            if productTile:
                theHREF = productTile[0].find_element_by_xpath('.//div[@class="image-panel"]/a').get_property('pathname').split('/').pop()
                beenDone = theHREF + siteDomain
                if beenDone not in replacedElems['prodID'] and replacedElems[elemRef]['initialColour']['ref']:
                    # Selects default initial product colour before any further interactions (if PLP)
                    # Can cause issues if different colour to that in recording
                    # IgnoreFlagDoNotInclude set as title to be picked up in Adobe server call
                    # Copy of mkorsData taken before click, which replaces the later version with the colour change click
                    selector = '.product-tile img[data-skuimg="' + replacedElems[elemRef]["initialColour"]["ref"] + '"]'
                    if 'selected' not in driver.find_element_by_css_selector(selector).find_element_by_xpath('..').get_attribute('class'):
                        driver.execute_script("swatch=document.querySelector('"+selector+"');swatch.title = 'IgnoreFlagDoNotInclude';temp=JSON.stringify(mkorsData);mkorsData='';swatch.click()")
                        time.sleep(1)
                        driver.execute_script("swatch.title='"+replacedElems[elemRef]['initialColour']['colour']+"';mkorsData=JSON.parse(temp)")
                        replacedElems['prodID'].append(theHREF + siteDomain)

        returnElem = newi or current
        print 'AFTER waitForElementAndSelect'
        print returnElem
        return [elem,returnElem,replacedElems,[oldURL,oldQB]]

    if '/checkout/' in driver.current_url:
        if i['target']=='[id="shippingAddress"]' and findCountry == 0:
            # Override any shipping address finder actions
            # Due to cross-country differences in postcode
            # Best practise is to manually enter address when recording, this is a fall back
            findCountry = 1
            autoScript = [
                ['.link-primary.link-Underlined','click()'],
                ['[id=address1]','send_keys("Test")'],
                ['[id=city]','send_keys("Test")'],
                ['[id=postalCode]','send_keys("'+countryPostal+'")']]
            conditionA = EC.presence_of_element_located((By.CSS_SELECTOR,autoScript[0][0]))
            elemA = WebDriverWait(driver, 10).until(conditionA)
            visibleClick(elemA,i)
            for j in autoScript[1:]:
                conditionA = EC.presence_of_element_located((By.CSS_SELECTOR,j[0]))
                elemA = WebDriverWait(driver, 10).until(conditionA)
                elemA.clear()
                exec('elemA.'+j[1]) in globals(), locals()

        if i['target'] == '[id="postalCode"]' and i['action']=='type':
            i['value'] = countryPostal

        if i['target'] == '[id="phone"]' and i['action']=='type':
            i['value'] = '000000000000'

        if i['target'] == 'label > div.row > div.col-xs-12 > p.card_details_text > span.paymentlabel':
            i['target']='[for=cc-payment]'

    if findCountry and not(i['target']=='[id="shippingAddress"]' or i['target'] == 'a.searchResultItem' or '//a[contains(text()' in i['target']):
        # Returns to regular script execution when no longer in shipping address finder
        findCountry = 0

    if not findCountry:

        elements = waitForElementAndSelect(i,replacedElems,oldURL,oldQB)
        elem = elements[0]
        newi = elements[1]
        replacedElems = elements[2]
        olds = elements[3]

        if newi['action']=='click':
            visibleClick(elem,newi)
        elif newi['action']=='submit':
            elem.submit()
        elif newi['action']=='type':
            elem.clear()
            elem.send_keys(newi['value'])
        elif newi['action']=='sendKeys':
            elem.send_keys(newi['value'])
        elif newi['action']=='mouseOver':
            hover = ActionChains(driver).move_to_element(elem)
            hover.perform()
        elif newi['action']=='select':
            # Drop down value given in recording is taken from element text
            # Here we search for the actual selector for this element
            value = re.match("(label=)?(.*)",newi["value"]).groups()[1]
            try:
                if driver.find_elements_by_css_selector('option[value="'+value+'"]'):
                    optionElem = value
                else:
                    optionElem = elem.find_element_by_xpath('//option[contains(text(), "'+value+'")]').get_attribute('value')
                driver.execute_script("document.querySelector('"+newi["target"]+"').value = '"+optionElem+"'")
            except WebDriverException:
                pass

        if 'customize=1' in driver.current_url:

            activeTab = "return document.querySelectorAll('.fc-ca-alias-colorway [role=tab][aria-expanded=true],.fc-ca-alias-monogram [role=tab][aria-expanded=true]').length"

            if newi['action']=='click' and 'mCSB_' in newi['target'] and not driver.execute_script(activeTab):
                # Removes keychains / straps from customisation if they are not in stock
                # (Checks for existence of price next to item description)
                typeAndPriceLabel = '.fc-nav-flyout-active .fc-attribute-selector-custom-before-text-hook'
                condition = EC.presence_of_element_located((By.CSS_SELECTOR,typeAndPriceLabel))
                elem = WebDriverWait(driver, 10).until(condition)

                if not re.search('[0-9]',elem.get_property('innerHTML')):
                    deSelectAddOn = '.fc-nav-flyout-active [aria-label="None None Selected"] span'
                    condition = EC.presence_of_element_located((By.CSS_SELECTOR,deSelectAddOn))
                    elem = WebDriverWait(driver, 10).until(condition)
                    elem.click()

            if newi['target']=="//div[@id='fluidConfigure']/div/div[2]/div[4]/div[4]/div/div/span":
                # Customised products take a long time to add to basket, so wait until it is present there before continuing
                condition = EC.presence_of_element_located((By.CSS_SELECTOR,'.mini-cart-list img[src*="'+driver.execute_script('return mkorsData.product.mfrItemNum')+'"]'))
                WebDriverWait(driver, 10).until(condition)


    return [olds[0],olds[1],replacedElems,findCountry]

def waitForAnalytics(oldURL,oldQB,newi,*last):
    # Waits until new analytics request has taken place on URL change
    # Explicit 3 second wait between actions that do not change page
    if oldURL != re.sub('#.*','',driver.current_url):
        oldURL = re.sub('#.*','',driver.current_url)
        waitScript = 'try{return window.top.s_gi(_satellite.sc_account).qb}catch(e){return 0}'
        newQB = driver.execute_script(waitScript)
        error = 10
        while error and (newQB==oldQB or not newQB):
            print 'waitForAnalytics wait'
            time.sleep(1)
            newQB = driver.execute_script(waitScript)
            error-=1
            if error == 0:
                debug.append('waitforAnalytics failed')
        time.sleep(1)
        oldQB = newQB
    elif newi['action']=='click' and 'checkout/checkout.jsp' not in driver.current_url:
        # Paused only during main site navigation, checkout steps do not trigger analytics calls
        time.sleep(3)
    return [oldURL,oldQB]

def visibleClick(elem,i):
    # Fall back to JavaScript click when event is not 'visible'
    # Also ensures element is top most in DOM i.e. no overlaying modals blocking click
    scrolling = 1
    clickJS = ''

    while scrolling:
        start = driver.execute_script("return window.scrollY")
        time.sleep(0.25)
        scrolling = 0 if start == driver.execute_script("return window.scrollY") else 1
        if scrolling:
            print 'INFO: Waiting - window scrolling'

    if not elem.is_displayed():
        if i['finder']=='xpath':
            clickJS = 'document.evaluate(\"'+i['target']+'\", document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.click()'
        elif i['finder']=='css':
            clickJS = 'document.querySelector(\''+i['target']+'\').click()'
        driver.execute_script(clickJS)

    error = 10
    while error and not clickJS:
        try:
            elem.click()
            error = 0
        except driverExceptions.WebDriverException, e:
            if 'Other element would receive the click' in e.msg:
                if error > 1:
                    print 'INFO: Waiting - other element overlaying target'
                    ActionChains(driver).move_to_element(elem).perform()
                    error -= 1
                    time.sleep(0.25)
                else:
                    if i['finder']=='xpath':
                        clickJS = 'document.evaluate(\"'+i['target']+'\", document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.click()'
                    elif i['finder']=='css':
                        clickJS = 'document.querySelector(\''+i['target']+'\').click()'
                    driver.execute_script(clickJS)
            else:
                raise e

def parseResponse(networkLog):
    # Transforms Chrome Driver log file into array format
    iter = 0
    responseObject = {}
    for i in networkLog:
        split = i.split('&')
        for j in split:
            try:
                keyValue = re.match('(^[^=]+)=(.*)',j).groups()
                key = keyValue[0].replace('c','_prop').replace('v','_evar') if re.match('(c|v)[0-9]',keyValue[0]) else keyValue[0]

                if re.match('(c|v)[0-9]',keyValue[0]):
                    key = keyValue[0].replace('c','_prop').replace('v','_evar')
                    if re.match('^(v1234|v9876|c12345)$',keyValue[0]): # To place individual props/evars higher up in output .xlsx
                        key = '_' + key
                elif re.match('^(AQB|AQE|D)$',keyValue[0]):
                    key = keyValue[0].lower()
                elif re.match('^(pev2|pageName|events|link)$',keyValue[0]): # To place other vars higher up in output .xlsx
                    key = '__'+keyValue[0]
                value = unquote(keyValue[1])
                try:
                    responseObject[key][iter]=value
                except KeyError:
                    responseObject[key] = {iter:value}
            except AttributeError:
                pass
        iter += 1
    # Breaks down individual products (attributes and events/evars) in response into product1,product2,...,productN
    for k in responseObject['products']:
        productNumber = 1
        for l in responseObject['products'][k].split(','):
            breakdown = re.match('^([^;]*);?([^;]*)?;?([^;]*)?;?([^;]*)?;?([^;]*)?;?(.*)',l).groups()
            try:
                responseObject['a_product '+str(productNumber)+' category'][k] = breakdown[0]
            except KeyError:
                responseObject['a_product '+str(productNumber)+' category'] = {k:breakdown[0]}
            try:
                responseObject['a_product '+str(productNumber)+' product'][k] = breakdown[1]
            except KeyError:
                responseObject['a_product '+str(productNumber)+' product'] = {k:breakdown[1]}
            try:
                responseObject['a_product '+str(productNumber)+' quantity'][k] = breakdown[2]
            except KeyError:
                responseObject['a_product '+str(productNumber)+' quantity'] = {k:breakdown[2]}
            try:
                responseObject['a_product '+str(productNumber)+' price'][k] = breakdown[3]
            except KeyError:
                responseObject['a_product '+str(productNumber)+' price'] = {k:breakdown[3]}

            if breakdown[4]!='':
                for m in breakdown[4].split('|'):
                    eventKeyValue = re.match('(^[^=]+)?=?(.*)',m).groups()
                    eventKey = eventKeyValue[0]
                    eventValue = eventKeyValue[1]
                    try:
                        responseObject['a_product '+str(productNumber)+' '+eventKey][k] = eventValue
                    except KeyError:
                        responseObject['a_product '+str(productNumber)+' '+eventKey] = {k:eventValue}
            if breakdown[5]!='':
                for n in breakdown[5].split('|'):
                    evarKeyValue = re.match('(^[^=]+)=?(.*)',n).groups()
                    evarKey = evarKeyValue[0]
                    evarValue = evarKeyValue[1]
                    try:
                        responseObject['a_product '+str(productNumber)+' '+evarKey][k] = evarValue
                    except KeyError:
                        responseObject['a_product '+str(productNumber)+' '+evarKey] = {k:evarValue}
            productNumber = productNumber + 1
    return [responseObject,iter]

def main(side,siteRef,env,device):

    parsedScript = parseSeleniumScript(side)
    # siteList must match exactly the array returned by tkWindow()
    siteList = ['co.uk','de/en_DE','fr/en_FR','it/en_IT','es/en_ES','ch/en_CH','eu/en_HU','eu/en_PL','eu/en_NL']
    global replacedElems
    replacedElems = {'prodID':[]}
    sites = []
    failed = []
    failCache = {}
    log = []
    global dataLayer
    dataLayer = {}

    for site in range(len(siteList)):
        if siteRef[site]:
            sites.append(siteList[site])
    """
    include = 'n' or raw_input('Generate expected Adobe values from prod? (Y/N): ')
    if len(include) and include.lower()[0]=='y':
        seleniumMain(parsedScript,'https://www.michaelkors.co.uk')
        included = 1
    else:
        included = 0"""

    for i in sites:
        try:
            seleniumMain(parsedScript,'https://'+env+'.michaelkors.'+i,replacedElems,sites,device)
            network = driver.get_log('performance')
            log += filterNetworkLog(network)
            layer = doDL('andReturn')
            dataLayer.update(layer)
        except KeyboardInterrupt:
            debug.append(i+' Keyboard Interrupt')
            time.sleep(5)
            network = driver.get_log('performance')
            failCache['_'+i] = filterNetworkLog(network)
            failed.append(i)
        except Exception,e:
            debug1 = '\nhttps://'+env+'.michaelkors.'+i+' FAILED! Will retry later\n'
            debug2 = str(type(e))
            debug3 = str(e)
            print debug1+'\n'+debug2+'\n'+debug3
            debug.append(debug1)
            debug.append(driver.current_url)
            debug.append(' :: '+debug2)
            debug.append(' :: '+debug3)
            time.sleep(5)
            network = driver.get_log('performance')
            failCache['_'+i] = filterNetworkLog(network)
            failed.append(i)

    for fails in failed:
        try:
            seleniumMain(parsedScript,'https://'+env+'.michaelkors.'+fails,replacedElems,sites,device)
            network = driver.get_log('performance')
            log += filterNetworkLog(network)
            failed.remove(fails)
            layer = doDL('andReturn')
            dataLayer.update(layer)
        except KeyboardInterrupt:
            debug.append(fails+' Keyboard Interrupt')
            time.sleep(5)
            network = driver.get_log('performance')
            failCache['_'+fails] = filterNetworkLog(network)
            log += failCache['_'+fails]
            layer = doDL('andReturn')
            dataLayer.update(layer)
        except Exception,e:
            debug1 = '\nhttps://'+env+'.michaelkors.'+fails+' FAILED TWICE!\n'
            debug2 = str(type(e))
            debug3 = str(e)
            print debug1+'\n'+debug2+'\n'+debug3
            debug.append(debug1)
            debug.append(driver.current_url)
            debug.append(' :: '+debug2)
            debug.append(' :: '+debug3)
            time.sleep(5)
            network = driver.get_log('performance')
            failCache['_'+fails] = filterNetworkLog(network)
            log += failCache['_'+fails]
            layer = doDL('andReturn')
            dataLayer.update(layer)

    fileDate = time.strftime('%Y-%m-%d %H.%M',time.localtime())
    fileSide = re.sub('.*(\/|\\\\)','',side).replace('.xml','')

    xlFile = 'G:\\eCommerce Europe\\BAU\\Analytics & Reporting\\6- Tools\\Selenium\\Results\\'+fileSide+' '+os.environ.get('USERNAME')+' '+fileDate+'.xlsx'
    debugFile = 'G:\\eCommerce Europe\\BAU\\Analytics & Reporting\\6- Tools\\Selenium\\Results\\logs\\'+fileSide+' '+os.environ.get('USERNAME')+' '+fileDate+'.txt'

    global parsed
    parsed = parseResponse(log)
    dataLayers = sortDataLayer(dataLayer,parsed[0]['t'])
    writeToXL(parsed[0],parsed[1],dataLayers,xlFile)
    doDebug(debugFile)

    #'Completed' dialogue box
    root = tk.Tk()
    root.withdraw()
    root.update_idletasks()
    root.title('Completed')
    root.resizable(0, 0)
    root.attributes("-toolwindow",1)
    x = (root.winfo_screenwidth() - root.winfo_reqwidth()) / 2
    y = (root.winfo_screenheight() - root.winfo_reqheight()) / 2
    root.geometry("+%d+%d" % (x, y))
    l1 = tk.Label(root, text="Completed with "+str(len(failed))+" / "+str(len(sites))+" site errors")
    l1.grid(row=0, sticky='N',pady=15,padx=20)
    button = tk.Button(root, text="Ok", command=root.destroy,width=10)
    button.grid(row=1,column=0,sticky='N',pady=10,padx=2)
    root.deiconify()
    root.mainloop()

    driver.quit()
    globals().pop('driver')

def doDebug(debugFile):
    #Writes debug file in /results/logs folder
    log = open(debugFile,'w')
    for i in debug:
        log.write(i+'\n')
    log.close()

def sortDataLayer(DLs,aaTimes):
    #Takes captured data layers, flattens and formats them for insertion in to .xlsx
    #Credit to https://stackoverflow.com/a/21108904 for _flatten_items  and flatten_dict

    def _flatten_items(items, sep, prefix):
        _items = []
        for key, value in items:
            _prefix = "{}{}".format(prefix, key)
            if isinstance(value, list):
                _items.extend(_flatten_items(list(enumerate(value)), sep=sep, prefix=_prefix+sep))
            elif isinstance(value, dict):
                _items.extend(_flatten_items(value.items(), sep=sep, prefix=_prefix+sep))
            else:
                _items.append((_prefix, value))
        return _items

    def flatten_dict(d, sep='_'):
        return dict(_flatten_items(d.items(), sep=sep, prefix=""))

    timeList = []
    sortedDL = sorted(DLs)
    idx = 0
    times = [('0'+re.search(' ([0-9][0-9]?):',aaTimes[x]).groups()[0])[-2:]+':'+('0'+re.search(':([0-9][0-9]?):',aaTimes[x]).groups()[0])[-2:]+':'+('0'+re.search(':([0-9][0-9]?) ',aaTimes[x]).groups()[0])[-2:] for x in aaTimes]

    for timesAA in times:
        while timesAA >= sortedDL[idx]:
            idx +=1
        timeList.append(sortedDL[idx])

    flatDLs = []
    for dlt in timeList:
        temp = flatten_dict(DLs[dlt]['dl'])
        last = []
        strList = []
        for i in sorted(temp):
            spl = i.split('_')
            indent = 0
            for j in spl:
                try:
                    lastIndex = last.index(j)
                except:
                    lastIndex = 'NotFound'
                if spl.index(j) != lastIndex:
                    if indent == len(spl)-1:
                        strList.append(indent * '  ' + '> ' + j+': ' + unicode(temp[i]))
                    else:
                        strList.append(indent * '  ' + '> ' + j+':')
                indent += 1
            last = spl
        flatDLs.append(strList)
    return flatDLs

def writeToXL(theResponse,iter,DLs,xlFile):
    #Writes datalayer and server calls to Excel

    fullNames = eventArray()
    wb = Workbook()
    wb.active.title = 'raw'
    sheets = [wb.active]
    cell = sheets[0]['A1']

    #Possibility that some server calls didn't fire, leading to inconsistencies in tabs
    #This aggregates all server calls across all countries
    print 'Processing Excel output: 10%'

    callsLog = {}
    tempArr = {}
    pageNameRef = {}
    for i in theResponse['g']:
        newUniqueCall = ''
        site = re.match('.*michaelkors([^/]+\/([a-zA-Z]{2}_[a-zA-Z]{2})?)',theResponse['g'][i]).groups()[0]

        try:
            pev = theResponse['__pev2'][i]
        except KeyError:
            pev = ''

        try:
            pageName = theResponse['__pageName'][i]
        except KeyError:
            pageName = ''

        theG = re.match('.*michaelkors([^/]+\/([a-zA-Z]{2}_[a-zA-Z]{2}\/?)?)([^?]+)?',theResponse['g'][i]).groups()[2] or ''
        if theG:
            splitG = theG.split('/')
            g0 = splitG.pop()
            g1 = splitG.pop()
            theG = g1+'/'+g0
        uniqueCall = theG + pev

        if '_'+site not in callsLog:
            callsLog['_'+site] = []
        if '_'+site not in tempArr:
            tempArr['_'+site] = []

        if uniqueCall in tempArr['_'+site]:
            newUniqueCall = uniqueCall+str(tempArr['_'+site].count(uniqueCall))
            callsLog['_'+site].append(newUniqueCall)
        else:
            callsLog['_'+site].append(uniqueCall)

        callToUse = newUniqueCall or uniqueCall
        pageNameRef['_'+callToUse] = {'pageName':pageName,'pev2':pev,'g':theResponse['g'][i]}
        tempArr['_'+site].append(uniqueCall)

    allCalls = []
    lastIndex = 0
    for x in callsLog:
        for y in callsLog[x]:
            try:
                index = allCalls.index(y)
            except ValueError:
                index = 'No index'
            if index=='No index':
                allCalls.insert(lastIndex+1,y)
                index = lastIndex + 1
            lastIndex = index

    #Generates new tab for each page call, names them by page name or event
    pagesArr = []
    for m in allCalls:
        if pageNameRef['_'+m]['pev2']:
            pageName = pageNameRef['_'+m]['pev2'][:30].replace(':','>')
        else:
            pageName = pageNameRef['_'+m]['pageName'].split('/').pop().split(' > ').pop()[:30].replace(':','>')
        if pageName in pagesArr:
            useName = pageName[:28]+'_'+str(pagesArr.count(pageName))
        else:
            useName = pageName
        pagesArr.append(pageName)
        sheets.append(wb.create_sheet(title=useName))

    #Populates 'raw' page

    for hosts in theResponse['g']:
        cell.offset(0,hosts+1).value = re.match('.*michaelkors([^/]+\/([a-zA-Z]{2}_[a-zA-Z]{2})?)',theResponse['g'][hosts]).groups()[0]
        cell.offset(0,hosts+1).font = Font(bold=True)
    row = 1
    for k in sorted(theResponse):

        try:
            if re.match('_prop[0-9]',k):
                cell.offset(row,0).value = fullNames['prop'+k[5:]]
            elif re.match('__prop[0-9]',k):
                cell.offset(row,0).value = fullNames['prop'+k[6:]]
            elif re.match('_evar[0-9]',k):
                cell.offset(row,0).value = fullNames['evar'+k[5:]]
            elif re.match('__evar[0-9]',k):
                cell.offset(row,0).value = fullNames['evar'+k[6:]]
            elif k[:10]=='a_product ':
                if ' e' in k:
                    cell.offset(row,0).value = re.sub(' [a-zA-Z].*',' ',k[1:]) + fullNames[re.sub('a_product [0-9][0-9]? ','',k.lower())]
                else:
                    cell.offset(row,0).value = k[1:]
            else:
                cell.offset(row,0).value = k.replace('__','')

        except KeyError:
            cell.offset(row,0).value = k.replace('__','')

        for l in range(iter):
            try:
                cell.offset(row,l+1).value = theResponse[k][l]
            except KeyError:
                cell.offset(row,l+1).value = '-'
        row = row + 1

    #Populates 'raw' page datalayer values
    dlCol = 0
    cell.offset(row + 2,0).value = 'mkorsData'
    cell.offset(row + 3,0).value = '(at time of server call)'
    cell.offset(row + 2,0).font = Font(name='Consolas', size=10)
    cell.offset(row + 3,0).font = Font(name='Consolas', size=10)

    for hitDL in DLs:
        dlRow = row + 2
        dlCol += 1
        for perLine in hitDL:
            cell.offset(dlRow,dlCol).value = perLine
            cell.offset(dlRow,dlCol).font = Font(name='Consolas', size=10)
            dlRow += 1
    print 'Processing Excel output: 20%'
    #Populates further tabs
    missing = 0
    for n in sheets[0][1][1:]:
        toMatch = 'toMatch'
        theMatch = 'theMatch'
        newIndex = ''
        while theMatch!=toMatch and newIndex!=0:
            toMatch = re.match('.*michaelkors([^/]+\/([a-zA-Z]{2}_[a-zA-Z]{2}\/?)?)([^?]+)?',theResponse['g'][(n.col_idx - 2)]).groups()[2] or ''
            if toMatch:
                splitMatch = toMatch.split('/')
                match0 = splitMatch.pop()
                match1 = splitMatch.pop()
                toMatch = match1+'/'+match0
            try:
                toMatch = toMatch + theResponse['__pev2'][(n.col_idx - 2)]
            except KeyError:
                toMatch = toMatch + '-'
            newIndex = ((n.col_idx - 2 + missing) % len(pageNameRef))
            theMatch = re.match('.*michaelkors([^/]+\/([a-zA-Z]{2}_[a-zA-Z]{2}\/?)?)([^?]+)?',pageNameRef['_'+allCalls[newIndex]]['g']).groups()[2] or ''

            if theMatch:
                splittheMatch = theMatch.split('/')
                thematch0 = splittheMatch.pop()
                thematch1 = splittheMatch.pop()
                theMatch = thematch1+'/'+thematch0
            theMatch = theMatch + (pageNameRef['_'+allCalls[newIndex]]['pev2'] or '-')

            if theMatch != toMatch:
                column_index = int(math.ceil(float((n.col_idx-1 + missing))/len(pageNameRef)))
                page_index = ((n.col_idx - 2 + missing) % len(pageNameRef)) + 1
                new_page_row_index = 1
                for m in sheets[0]['A']:
                    if m.value != '-':
                        sheets[page_index].cell(new_page_row_index,column_index,"-")
                        new_page_row_index = new_page_row_index + 1
                missing += 1

        column_index = int(math.ceil(float((n.col_idx-1+missing))/len(pageNameRef)))
        page_index = ((n.col_idx - 2 + missing) % len(pageNameRef)) + 1
        new_page_row_index = 1

        for m in sheets[0]['A']:
            if m.value != '-':
                sheets[page_index].cell(new_page_row_index,column_index,sheets[0].cell(m.row,n.col_idx).value)
                sheets[page_index].cell(new_page_row_index,column_index).font = Font(name=sheets[0].cell(m.row,n.col_idx).font.name, size = sheets[0].cell(m.row,n.col_idx).font.size)
                new_page_row_index = new_page_row_index + 1

    #Deletes rows in further tabs with no values set in any calls
    percentCount = 0
    for sheet in sheets[1:]:
        percentage = 35 + ((sheets[1:].index(sheet))/float(len(sheets[1:])))*60
        print 'Processing Excel output: '+str(percentage)[:2]+'%'
        toDelete=[]
        sheet.insert_cols(0,1)
        for eachRow in sheet['A']:
            noneSet = 1
            for rowCell in sheet[eachRow.row][1:]:
                noneSet = 0 if rowCell.value != '-' else noneSet
            if noneSet:
                toDelete.append(eachRow.row)
            else:
                eachRow.value = sheets[0]['A'][eachRow.row-1].value
        toDelete.sort()
        toDelete.reverse()
        for deletable in toDelete:
            sheet.delete_rows(deletable,1)
        sheet.column_dimensions['A'].width = 42
        for otherCols in sheet[1][1:]:
            sheet.column_dimensions[otherCols.column].width = 28

        percentCount+=1

    for sheet in sheets:
        sheet.freeze_panes = "A2"
        for headers in sheet[1]:
            headers.font = Font(bold=True)
    for allWs in sheets:
        allWs.sheet_view.zoomScale = 80

    print 'Processing Excel output: 99%'
    wb.save(filename = xlFile)
    print 'File created successfully'
    os.startfile(xlFile)

def eventArray():
    # Add new props/evars/events here to these arrays in the same format, comma separated e.g.
    # ,"evarXX":"Descriptive Name"
    events = {"event1":"Internal Keyword Searches (e1)","event2":"Internal Keyword Null Searches (e2)","event3":"Product Views (Custom) (e3)","event4":"Bounce Rate - Entries (e4)","event5":"Bounce Rate - Click Past (e5)","event6":"Merch Instances (Custom) (e6)","event7":"Login - Checkout Step 1 (e7)","event8":"Shipping - Checkout Step 2 (e8)","event9":"Billing - Checkout Step 3 (e9)","event10":"Order Review - Checkout Step 4 (e10)","event11":"Discount (Product Level) (e11)","event12":"Discount (Order Level) (e12)","event13":"Tax (e13)","event14":"Shipping (e14)","event15":"Gift Card Used (e15)","event16":"Markdown Amount (e16)","event17":"Gift Wrap (e17)","event18":"Type Ahead Used (e18)","event19":"Recommended Search (e19)","event20":"Search Refinements (e20)","event21":"Quick Views (e21)","event22":"Product Detail Interactions (e22)","event23":"Find in Store (e23)","event24":"Favorites (e24)","event25":"Social Shares (e25)","event26":"Email Sign Up Starts (e26)","event27":"Email Sign Up Completes (e27)","event28":"Registrations (e28)","event29":"Video Starts (e29)","event30":"Video Completes (e30)","event31":"Slideshow Starts (e31)","event32":"Slideshow Completes (e32)","event33":"Pagination (e33)","event34":"Store Locator (e34)","event35":"UnFavorite (e35)","event36":"Desktop Site Clicks (e36)","event37":"Click to Call (e37)","event38":"Nearest Store Selection (e38) Custom 38","event39":"Directions (e39)","event40":"About KORS VIP (e40)","event41":"VIP Sign UP (e41)","event42":"Content Detail Interaction (e42)","event43":"Video 10% Watched (e43)","event44":"Video 20% Watched (e44)","event45":"Video 30% Watched (e45)","event46":"Video 40% Watched (e46)","event47":"Video 50% Watched (e47)","event48":"Video 60% Watched (e48)","event49":"Video 70% Watched (e49)","event50":"Video 80% Watched (e50)","event51":"Video 90% Watched (e51)","event52":"Gift Receipt (e52)","event53":"MGM Product View (e53)","event54":"MGM Product Configure (e54)","event55":"MGM Product Add to Cart (e55)","event56":"MGM Product Purchase (e56)","event57":"Size Selection (e57)","event58":"Out of Stock SKU selected (e58)","event59":"MGM Product Revenue (e59)","event60":"Markdown Revenue $ (e60)","event61":"Markdown Units (e61)","event62":"Email Opt Out Starts (e62)","event63":"Email Opt Out Completes (e63)","event64":"Product Selling Price (e64)","event65":"Cookie Policy Appears (e65)","event66":"Cookie Policy clicked (e66)","event67":"Country Selector (e67)","event68":"Country Selector Clicked (e68)","event69":"Size and Fit Guide Clicked (e69)","event70":"Size and Fit Guide Closed (e70)","event71":"Address Look-Up (e71)","event72":"Language Preference (e72)","event73":"Basket Porting (e73)","event74":"Find Address Checkout (e74)","event75":"MD is Present (e75)","event76":"Product Listing Page Interaction (e76)","event77":"VIP Opt Out (e77)","event78":"Custom Product View (e78)","event79":"Custom Configure View (e79)","event80":"Custom Add to Bag (e80)","event81":"Custom Purchase (e81)","event82":"Custom Revenue (e82)","event83":"Product Not Available (e83)","event84":"Number of Search Results (e84)","event85":"VIP Point Average (e85)","event86":"VIP Sign In (e86)","event87":"True Fit- Size displayed (e87)","event88":"True Fit- Find truefit interaction (e88)","event89":"Internal Search Initiation (e89)","event90":"Units Added to Cart (e90)","event91":"Units Removed from Cart (e91)","event99":"PDP Error (e99)","event100":"RR Product Impressions (e100)","event101":"RR Product Placement Index (e101)","event102":"GWP PDP Impression (e102)","event103":"GWP PDP Learn More | Alt img Click (e103)","event104":"Apple Pay (e104)","event105":"GWP Purchased Units (e105)","event106":"Out Of Stock (e106)","event111":"Emails Sent (e111)","event112":"Emails Delivered (e112)","event113":"Emails Opened (e113)","event114":"Emails Clicked (e114)","event115":"Emails Unsubscribed (e115)","event116":"Emails Bounced (e116)","event120":"Live Chat - Concierge Drawer (e120)","event121":"Live Chat - Knowledge (e121)","event122":"Live Chat - Chat (e122)","event123":"Live Chat - CoBrowse (e123)","event201":"Revenue USD (e201)","event202":"Revenue CAD (e202)","event203":"Revenue GBP (e203)","event204":"Revenue CHF (e204)","event205":"Revenue EUR (e205)","event211":"Discount - Product Level USD (e211)","event212":"Discount - Product Level CAD (e212)","event213":"Discount - Product Level GBP (e213)","event214":"Discount - Product Level CHF (e214)","event215":"Discount - Product Level EUR (e215)","event221":"Discount - Order Level USD (e221)","event222":"Discount - Order Level CAD (e222)","event223":"Discount - Order Level GBP (e223)","event224":"Discount - Order Level CHF (e224)","event225":"Discount - Order Level EUR (e225)","event231":"Tax USD (e231)","event232":"Tax CAD (e232)","event233":"Tax GBP (e233)","event234":"Tax CHF (e234)","event235":"Tax EUR (e235)","event241":"Shipping USD (e241)","event242":"Shipping CAD (e242)","event243":"Shipping GBP (e243)","event244":"Shipping CHF (e244)","event245":"Shipping EUR (e245)","event251":"Markdown Amount USD (e251)","event252":"Markdown Amount CAD (e252)","event253":"Markdown Amount GBP (e253)","event254":"Markdown Amount CHF (e254)","event255":"Markdown Amount EUR (e255)","event261":"Gift Wrap USD (e261)","event262":"Gift Wrap CAD (e262)","event263":"Gift Wrap GBP (e263)","event264":"Gift Wrap CHF (e264)","event265":"Gift Wrap EUR (e265)","event271":"Markdown Revenue USD (e271)","event272":"Markdown Revenue CAD (e272)","event273":"Markdown Revenue GBP (e273)","event274":"Markdown Revenue CHF (e274)","event275":"Markdown Revenue EUR (e275)","event281":"Product Selling Price USD (e281)","event282":"Product Selling Price CAD (e282)","event283":"Product Selling Price GBP (e283)","event284":"Product Selling Price CHF (e284)","event285":"Product Selling Price EUR (e285)","event291":"Monogram Revenue USD (e291)","event292":"Monogram Revenue CAD (e292)","event293":"Monogram Revenue GBP (e293)","event294":"Monogram Revenue CHF (e294)","event295":"Monogram Revenue EUR (e295)","event300":"BV Number of Reviews (e300)","event301":"BV Review Instances (e301)","event302":"BV AVG. Rating (e302)","event303":"BV Fit Score (e303)","event304":"BV Fashionable Score (e304)","event305":"BV Width Score (e305)","event306":"BV Fit at the Calf Score (e306)","event311":"BV Star Rating 1 (e311)","event312":"BV Star Rating 2 (e312)","event313":"BV Star Rating 3 (e313)","event314":"BV Star Rating 4 (e314)","event315":"BV Star Rating 5 (e315)","event351":"Custom Kors Revenue USD (e351)","event352":"Custom Kors Revenue CAD (e352)","event353":"Custom Kors Revenue GBP (e353)","event354":"Custom Kors Revenue CHF (e354)","event355":"Custom Kors Revenue EUR (e355)","event356":"Custom Kors Revenue CNY (e356)","event357":"Custom Kors Revenue JPY (e357)","event358":"Custom Kors Revenue ZAR (e358)","event359":"Custom Kors Revenue HKD (e359)","event401":"Revenue SEK (e401)","event402":"Revenue PLN (e402)","event403":"Revenue NOK (e403)","event404":"Revenue DKK (e404)","event405":"Revenue CZK (e405)","event406":"Revenue HUF (e406)","event411":"Discount - Product Level SEK (e411)","event412":"Discount - Product Level PLN (e412)","event413":"Discount - Product Level NOK (e413)","event414":"Discount - Product Level DKK (e414)","event415":"Discount - Product Level CZK (e415)","event416":"Discount - Product Level HUF (e416)","event421":"Discount - Order Level SEK (e421)","event422":"Discount - Order Level PLN (e422)","event423":"Discount - Order Level NOK (e423)","event424":"Discount - Order Level DKK (e424)","event425":"Discount - Order Level CZK (e425)","event426":"Discount - Order Level HUF (e426)","event431":"Tax SEK (e431)","event432":"Tax PLN (e432)","event433":"Tax NOK (e433)","event434":"Tax DKK (e434)","event435":"Tax CZK (e435)","event436":"Tax HUF (e436)","event441":"Shipping SEK (e441)","event442":"Shipping PLN (e442)","event443":"Shipping NOK (e443)","event444":"Shipping DKK (e444)","event445":"Shipping CZK (e445)","event446":"Shipping HUF (e446)","event451":"Markdown Amount SEK (e451)","event452":"Markdown Amount PLN (e452)","event453":"Markdown Amount NOK (e453)","event454":"Markdown Amount DKK (e454)","event455":"Markdown Amount CZK (e455)","event456":"Markdown Amount HUF (e456)","event461":"Gift Wrap SEK (e461)","event462":"Gift Wrap PLN (e462)","event463":"Gift Wrap NOK (e463)","event464":"Gift Wrap DKK (e464)","event465":"Gift Wrap CZK (e465)","event466":"Gift Wrap HUF (e466)","event471":"Markdown Revenue SEK (e471)","event472":"Markdown Revenue PLN (e472)","event473":"Markdown Revenue NOK (e473)","event474":"Markdown Revenue DKK (e474)","event475":"Markdown Revenue CZK (e475)","event476":"Markdown Revenue HUF (e476)","event481":"Product Selling Price SEK (e481)","event482":"Product Selling Price PLN (e482)","event483":"Product Selling Price NOK (e483)","event484":"Product Selling Price DKK (e484)","event485":"Product Selling Price CZK (e485)","event486":"Product Selling Price HUF (e486)","event491":"Monogram Revenue SEK (e491)","event492":"Monogram Revenue PLN (e492)","event493":"Monogram Revenue NOK (e493)","event494":"Monogram Revenue DKK (e494)","event495":"Monogram Revenue CZK (e495)","event496":"Monogram Revenue HUF (e496)","event501":"addToCart no scOpen - diagnostic","event502":"data_capture_form - diagnostic","event503":"changeSelProductOpt - diagnostic","event522":"Pick Up In Store Button Clicks (e522)","event523":"Add to Bag & Pickup in Store Attempt (e523)","event551":"Custom Kors Revenue SEK (e551)","event552":"Custom Kors Revenue PLN (e552)","event553":"Custom Kors Revenue NOK (e553)","event554":"Custom Kors Revenue DKK (e554)","event555":"Custom Kors Revenue CZK (e555)","event556":"Custom Kors Revenue HUF (e556)"}
    propsandvars = {"evar1":"Product Finding Methods (v1)","evar2":"Internal Search Keywords (v2)","evar3":"Internal Campaign Tracking Codes (v3)","evar4":"Merchandising Category Level 1 (v4)","evar5":"Merchandising Category Level 2 (v5)","evar6":"Merchandising Category Level 3 (v6)","evar7":"Cross-selling Pages (v7)","evar8":"Cross-selling Products (v8)","evar9":"Add-to-Cart Locations (v9)","evar10":"New/Repeat Visitors (v10)","evar11":"Order IDs (v11)","evar12":"Payment Methods (v12)","evar13":"Shipping Methods (v13)","evar14":"Time Parting (v14)","evar15":"Discount Codes (v15)","evar16":"Shipping State (v16)","evar17":"Shipping Zip Code (v17)","evar18":"Internal Campaign Origin (v18)","evar19":"Add-to-Cart Type (v19)","evar20":"Member/Non-Member (v20)","evar21":"User ID (v21)","evar22":"Omniture ID (v22)","evar23":"Refinement Value (v23)","evar24":"CRM Segment (v24)","evar25":"Product Detail Interaction Types (v25)","evar26":"Social Share Type (v26)","evar27":"Outfit ID (v27)","evar28":"Site Type (v28)","evar29":"Country/Language (v29)","evar30":"Checkout Flow (v30)","evar31":"Billing Country (v31)","evar32":"Shipping Country (v32)","evar33":"Nearest Store Selection (v33)","evar34":"Slideshow Name (v34)","evar35":"URL / Breadcrumb (v35)","evar36":"Add-to-Cart Site (v36)","evar37":"Article ID (v37)","evar38":"Find in Store Zip Code (v38)","evar39":"Article Name (v39)","evar40":"Video Name (v40)","evar41":"Module Type (v41)","evar42":"Email Sign Up Type (v42)","evar43":"ATG Category ID (v43)","evar44":"Shared Content (v44)","evar45":"Content Detail Interaction Type (v45)","evar46":"iOS Device Type (v46)","evar47":"Recipient ID (email) (v47)","evar48":"Message ID (v48)","evar49":"Marketing Email Type (v49)","evar50":"Email Address (v50)","evar51":"Previous Page URL (v51)","evar52":"Store Style # (v52)","evar53":"Basket Porting (v53)","evar54":"Adobe MID (v54)","evar55":"Country Selector (v55)","evar56":"Add to Cart Shipping Method (v56)","evar57":"Pick Up Store ID (v57)","evar58":"GWP Product Detail Interaction types (v58)","evar59":"Currency (v59)","evar60":"Billing Title (v60)","evar61":"Screen Orientation (v61)","evar62":"RR Type Text (v62)","evar63":"MD or FP (v63)","evar64":"BV Rating (v64)","evar65":"Monetate Campaign (v65)","evar66":"LookID (v66)","evar67":"Page Name (v67)","evar68":"Internal Search Type (v68)","evar69":"Product Listing Interaction Types (v69)","evar70":"ATG Category ID by Cart (v70)","evar71":"VIP Tier (v71)","evar72":"About KORS VIP (v72)","evar73":"VIP Sign Up Location (v73)","evar74":"VIP Points Action (v74)","evar75":"VIP Profile Type (v75)","evar90":"Basket Porting - Add to Cart (v90)","evar91":"Basket Porting - Purchase (v91)","evar92":"Reserved Live Chat v92","evar98":"GWP Sku (v98)","evar99":"RR Product Placement (v99)","evar100":"RR Products (v100)","evar101":"RR Products Positioning (v101)","evar102":"Live Chat - Interaction Type (v102)","evar103":"User Agent (v103)","evar104":"X-Purpose Header (v104)","evar105":"Test - Error (v105)","prop1":"Page Types (p1)","prop2":"Site Sections Level 2 (p2)","prop3":"Site Sections Level 3 (p3)","prop4":"Internal Search Keywords (p4)","prop5":"# of Search Results (p5)","prop6":"Previous Page (p6)","prop7":"Max Vertical Pixels of Page Viewed (p7)","prop8":"Keyword Search Results Tab (p8)","prop9":"Error Message (p9)","prop10":"Video Name (p10)","prop11":"Slideshow Name (p11)","prop12":"Article Name (p12)","prop13":"Module Type (p13)","prop14":"URL / Breadcrumb (p14)","prop15":"Location Services On (p15)","prop16":"Link Location (p16)","prop17":"iOS Device Type (p17)","prop18":"Previous Page Type (p18)","prop19":"Previous Page URL (p19)","prop20":"Store Style # (p20)","prop21":"User-Agent (p21)","prop22":"Search Result Text (p22)","prop23":"Search Results Filter (p23)","prop24":"Search Filter No. Results (p24)","prop25":"Experiment Test Tracker (p25)","prop26":"Category ID (p26)","prop27":"% of Page Viewed (p27)","prop74":"CrossSell Product x Index DEBUG (p74)","prop75":"DTM publish date (p75)"}
    events.update(propsandvars)
    return events

def tkWindow():
    # GUI window shown on script execution, passes variables to main code

    def get_side():
        name = tkFileDialog.askopenfilename(initialdir = "G:/eCommerce Europe/BAU/Analytics & Reporting/6- Tools/Selenium",title = "Select file",filetypes = (("Selenium IDE Script","*.xml"),("All Files","*.*")))
        entry_text.set(name)

    def doMain():
        # Pulls path to Katalon .xml script from entry box
        side = entry_text.get()

        siteRef = []
        environArr=[]
        devArr = []

        for j in checked:
            siteRef.append(j.get())

        for k in environs:
            environArr.append(k.get())

        for l in devices:
            devArr.append(l.get())


        if environArr.index(1)==0:
            environment = 'uat1'
        elif environArr.index(1)==1:
            environment = 'sit'
        elif environArr.index(1)==2:
            environment = 'www'

        if devArr.index(1)==0:
            device = 'desktop'
        elif devArr.index(1)==1:
            device = 'mobile'
        elif devArr.index(1)==2:
            device = 'tablet'
        root.destroy()
        main(side,siteRef,environment,device)

    def unCheck(n,row):
        toCheck = [0,1,2]
        toCheck.pop(n)
        for checks in toCheck:
            row[checks].set(0)


    root = tk.Tk()
    root.withdraw()
    root.update_idletasks()
    root.title('Analytics Gathering - Michael Kors UAT')
    root.resizable(0, 0)

    menubar = tk.Menu(root)

    filemenu = tk.Menu(menubar, tearoff=0)
    filemenu.add_command(label="Exit", command=root.destroy)
    menubar.add_cascade(label="File", menu=filemenu)
    root.config(menu=menubar)

    l1 = tk.Label(root, text="Select UAT sites (Default: All)")
    l1.grid(row=0, sticky='W',columnspan=9,pady=10)

    checked=[]
    labels = ['UK','DE','FR','IT','ES','CH','HU','PL','NL']

    for i in range(len(labels)):

        checked.append(tk.IntVar())

        check = tk.Checkbutton(root, text=labels[i], variable=checked[i])
        check.grid(row=1,column=i, padx=2)
        checked[i].set(1)

    l3 = tk.Label(root, text="Select environment")
    l3.grid(row=2, sticky='W',columnspan=9,pady=10)

    environs = [tk.IntVar(),tk.IntVar(),tk.IntVar()]

    en1 = tk.Checkbutton(root, text='UAT', variable=environs[0],command=lambda:unCheck(0,environs))
    en1.grid(row=3,column=0, padx=2,columnspan=3)
    en2 = tk.Checkbutton(root, text='SIT', variable=environs[1],command=lambda:unCheck(1,environs))
    en2.grid(row=3,column=3, padx=2,columnspan=3)
    en3 = tk.Checkbutton(root, text='WWW', variable=environs[2],command=lambda:unCheck(2,environs))
    en3.grid(row=3,column=6, padx=2,columnspan=3)

    environs[0].set(1)

    l4 = tk.Label(root, text="Select device")
    l4.grid(row=4, sticky='W',columnspan=9,pady=10)

    devices = [tk.IntVar(),tk.IntVar(),tk.IntVar()]

    dv1 = tk.Checkbutton(root, text='Desktop', variable=devices[0],command=lambda:unCheck(0,devices))
    dv1.grid(row=5,column=0, padx=2,columnspan=3)
    dv2 = tk.Checkbutton(root, text='Mobile', variable=devices[1],command=lambda:unCheck(1,devices))
    dv2.grid(row=5,column=3, padx=2,columnspan=3)
    dv3 = tk.Checkbutton(root, text='Tablet', variable=devices[2],command=lambda:unCheck(2,devices))
    dv3.grid(row=5,column=6, padx=2,columnspan=3)

    devices[0].set(1)

    l2 = tk.Label(root, text="Select .xml file")
    l2.grid(row=6, sticky='W',columnspan=9,pady=10)

    button = tk.Button(root, text="Browse...", command=get_side)
    button.grid(row=7,column=0,sticky='W',pady=5,padx=2,columnspan=3)

    entry_text = tk.StringVar()
    e1 = tk.Entry(root, textvariable=entry_text,width=50)
    e1.grid(row=7,column=2,sticky='W',pady=5,padx=2,columnspan=6)

    button1 = tk.Button(root, text="Execute", command=doMain,width=15)
    button1.grid(pady=10,columnspan=9)

    x = (root.winfo_screenwidth() - root.winfo_reqwidth()) / 2
    y = (root.winfo_screenheight() - root.winfo_reqheight()) / 2
    root.geometry("+%d+%d" % (x, y))
    root.deiconify()
    root.mainloop()

if __name__ == "__main__":
    tkWindow()